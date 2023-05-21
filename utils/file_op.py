'''
Desc:
File: /file_op.py
Project: utils
File Created: Sunday, 27th June 2021 12:29:28 am
Author: luxuemin2108@gmail.com
-----
Copyright (c) 2021 Camel Lu
'''

import time
from datetime import datetime
import os

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# 写json文件
def write_fund_json_data(data, filename, file_dir=None):
    import json
    if not file_dir:
        cur_date = time.strftime("%Y-%m-%d", time.localtime(time.time()))
        file_dir = os.getcwd() + '/archive_data/json/' + cur_date + '/'
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
        print("目录新建成功：%s" % file_dir)
    with open(file_dir + filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.close()


def update_xlsx_file(path, df_data, sheet_name, *, index=False):
    if os.path.exists(path):
        writer = pd.ExcelWriter(path, engine='openpyxl')
        book = load_workbook(path)
        # 表名重复，删掉，重写
        if sheet_name in book.sheetnames:
            del book[sheet_name]
        if len(book.sheetnames) == 0:
            df_data.to_excel(
                path, sheet_name=sheet_name, index=index)
            return
        else:
            writer.book = book
            df_data.to_excel(
                writer, sheet_name=sheet_name, index=index)
        writer.save()
        writer.close()
    else:
        df_data.to_excel(
            path, sheet_name=sheet_name, index=index)

def read_bond_excel(*, source_date=datetime.now().strftime("%Y-%m-%d"), sheetname="多因子", is_bond=True):
    path = '/Users/admin/personal/anchor_plan/convertible-bond-crawler/dynamic_out/' + \
        source_date + '_cb_list.xlsx'
    xls = pd.read_excel(path, dtype={
            "可转债代码": np.str, 
            "股票代码": np.str, 
            }, engine='openpyxl', sheet_name=None)
    df_multi_factors = xls[sheetname]
    columns={
        '可转债代码': 'code',
        '可转债名称': 'name',
        '股票代码': 'stock_code',
        '股票名称': 'stock_name',
        '市场': 'market',
    }
    if not is_bond:
        columns={
            '可转债代码': 'bond_code',
            '可转债名称': 'bond_name',
            '股票代码': 'code',
            '股票名称': 'name',
            '市场': 'market',
        }
    df_multi_factors = df_multi_factors.rename(columns=columns).reset_index(drop=True)
    df_multi_factors['code'] = df_multi_factors['code'].astype(str)
    df_multi_factors['symbol'] = df_multi_factors['market'] + \
        df_multi_factors['code']
    return df_multi_factors
